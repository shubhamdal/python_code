#!/usr/bin/python
from pyspark.sql import SparkSession
from pyspark.sql.types import *
from pyspark.sql.functions import *
import sys
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import json
import pyspark.sql.functions as func

def load(paths,paths_dic,spark):
    paths_list = paths.split(',')
    for i in range(len(paths_list)):
        paths_dic[i] = spark.read.format("csv").load(paths_list[i],header='true',inferSchema='true')
    return paths_dic

def join(tbl1, tbl2, type, columns, join_select_cols):
    if join_select_cols[0].lower() == 'all':
        df = tbl1.alias('t1').join(tbl2.alias('t2'), on=columns, how=type.lower())
    else:
        df = tbl1.alias('t1').join(tbl2.alias('t2'), on=columns, how=type.lower()).select(join_select_cols)
    return df

def select(df, cols):
    column_list = cols.split(',')
    df1 = df.select([col for col in df.columns if col in column_list])
    return df1

def filter(df, conditions):
    df1 = df.filter(conditions)
    return df1

def orderby(df, columns, type):
    if type.lower() == 'desc':
        df1 = df.orderBy(df[columns].desc())
    else:
        df1 = df.orderBy(df[columns].asc())
    return df1

def groupby(df, columns):
    df1 = df.groupBy(columns)
    return df1

def count(df, column, group_flag):
    if group_flag.lower() == 'true':
        df1 = df.count()
    else:
        df1 = df.groupBy(column).count()
    return df1

def sql(df, sql_query,spark):
    df.createOrReplaceTempView("df")
    df1 = spark.sql(sql_query)
    return df1

def save(df, type, delimiter, path):
    if type.lower() == 'text' or type.lower() == 'csv':
        df.coalesce(1).write.option("sep",delimiter).option("header","true").csv(path)
    elif type.lower() == 'excel':
        df_pd = df.toPandas()
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            workbook=openpyxl.Workbook()
            writer.book=workbook
            writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
            print(writer.sheets)
            worksheet_names = workbook.sheetnames
            sheet_index = worksheet_names.index('Sheet')
            print("sheet index: "+(str(sheet_index)))
            workbook.active = sheet_index
            df_pd.to_excel(writer,sheet_name='Sheet',startrow=0,index=False, header=True)
            writer.save()
    print("Report saved successfully...")


def main(parameter_file):
    spark = SparkSession.builder.appName("Assignment_Ankit").enableHiveSupport().getOrCreate()
    paths_dic = {}
    group_flag = 'false'
    with open(parameter_file) as param:
        steps = json.load(param)
        for i in range(len(steps)):
            j = i + 1
            for step in steps:
                if str(j)+"_" in step:
                    print(step)
                    if "load" in step.lower():
                        print("Loading Dataframe..")
                        load_step = steps[step]
                        load_step_list = load_step.split(';')
                        path_str = load_step_list[0].strip("(").strip(")")
                        paths_dic = load(path_str,paths_dic,spark)
                    elif "join" in step.lower():
                        print("Joining Dataframe..")
                        join_step = steps[step]
                        join_step_list = join_step.split(';')
                        tbl1 = paths_dic[0]
                        tbl2 = paths_dic[1]
                        join_type = join_step_list[0].strip("(").strip(")")
                        join_cols = (join_step_list[1].strip("(").strip(")")).split(',')
                        join_select_cols = (join_step_list[2].strip("(").strip(")")).split('=')[1]
                        join_select_cols_list = join_select_cols.split(',')
                        outdf = join(tbl1,tbl2,join_type,join_cols,join_select_cols_list)
                    elif "select" in step.lower():
                        print("Selecting Dataframe..")
                        select_step = steps[step]
                        select_step_list = select_step.split(';')
                        cols = select_step_list[0].strip("(").strip(")")
                        if j == 2:
                            outdf = select(paths_dic[0], cols)
                        else:
                            outdf = select(outdf, cols)
                    elif "filter" in step.lower():
                        print("Filtering Dataframe..")
                        filter_step = steps[step]
                        filter_step_list = filter_step.split(';')
                        filter_conditions = filter_step_list[0].strip("(").strip(")")
                        if j == 2:
                            outdf = filter(paths_dic[0], filter_conditions)
                        else:
                            outdf = filter(outdf, filter_conditions)
                    elif "groupby" in step.lower():
                        print("Grouping Dataframe..")
                        group_flag = 'true'
                        group_step = steps[step]
                        group_step_list = group_step.split(';')
                        group_col = (group_step_list[0].strip("(").strip(")")).split(',')
                        if j == 2:
                            outdf = groupby(paths_dic[0],group_col)
                        else:
                            outdf = groupby(outdf,group_col)
                    elif "count" in step.lower():
                        print("Counting Dataframe..")
                        print("group_flag:"+group_flag)
                        count_step = steps[step]
                        count_step_list = count_step.split(';')
                        count_col = (count_step_list[0].strip("(").strip(")")).split(',')
                        if j == 2:
                            outdf = count(paths_dic[0],count_col,group_flag)
                        else:
                            outdf = count(outdf,count_col,group_flag)
                    elif "orderby" in step.lower():
                        print("Ordering Dataframe..")
                        order_step = steps[step]
                        order_step_list = order_step.split(';')
                        order_cols = order_step_list[0].strip("(").strip(")")
                        order_type = order_step_list[1].strip("(").strip(")")
                        if j == 2:
                            outdf = orderby(paths_dic[0],order_cols,order_type)
                        else:
                            outdf = orderby(outdf,order_cols,order_type)
                    elif "sql" in step.lower():
                        print("Running SQL on Dataframe..")
                        sql_step = steps[step]
                        sql_step_list = sql_step.split(';')
                        sql_query = sql_step_list[0]
                        if j == 2:
                            outdf = sql(paths_dic[0],sql_query,spark)
                        else:
                            outdf = sql(outdf,sql_query,spark)
                    elif "save" in step.lower():
                        print("Saving Report..")
                        save_step = steps[step]
                        save_step_list = save_step.split(';')
                        file_type = save_step_list[0].strip("(").strip(")")
                        delimiter = save_step_list[1].strip("(").strip(")")
                        outpath = save_step_list[2].strip("(").strip(")")
                        if j == 2:
                            save(paths_dic[0], file_type, delimiter, outpath)
                        else:
                            save(outdf, file_type, delimiter, outpath)
    spark.stop()

if __name__ == '__main__':
    if len(sys.argv) < 1:
        sys.exit("Parameter json file not passed as an argument..")
    parameter_file = sys.argv[1]
    main(parameter_file)
