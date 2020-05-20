import os
import pandas as pd
from query_service import SQLFailedError
from query_service import exec_teradata_sql
import openpyxl
from openpyxl import load_workbook
from collections import OrderedDict
from query_service import *
import re
from collections import OrderedDict
import traceback
import sys
from datetime import datetime
import pytz
from dateutil.relativedelta import relativedelta


# sys.path.insert(0,'/Users/shuchawla/documents/new_airflow/MR_Jobs/credentials')
sys.path.insert(1, '/home/sburman4/MR_Jobs/credentials')
import cred


def run_query_file(infile,outfile,comments_od,verbose=False):
    with open(infile, 'r') as f:
        queries = [q.strip() for q in re.split(r';[\s\r\n]+', f.read()) if q.strip()]
        for i,query in enumerate(queries):
            query_list=re.split(r'-->', str(query))
            query_number=query_list[0]
            query_string=query_list[1]
            query_comment=comments_od[str(query_number)]        
#             print("query_comment="+query_comment)
            if verbose:
                print('Running query: {}/{}\n{}'.format(i+1, len(queries), q))
            try:
                columns,rows = exec_teradata_sql(query_string)
                df = pd.DataFrame(list(rows), columns=columns)
                df.columns = df.columns.str.upper()
                if not columns:
                    print('No output result')
                else:
                    print("saving in excel")
                    with pd.ExcelWriter(outfile, engine='openpyxl') as writer:
                        workbook=openpyxl.load_workbook(outfile)
                        workbook.create_sheet(query_number)
                        writer.book=workbook
                        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)   
                        print(writer.sheets)
                        sheet =workbook[query_number]
                        worksheet_names = workbook.sheetnames
                        sheet_index = worksheet_names.index(query_number)
                        print("sheet index: "+(str(sheet_index)))
                        workbook.active = sheet_index
                        df.to_excel(writer, query_number,startrow=7,index=False, header=True)
                        writer.save() 
                        workbook.save(outfile)
                    workbook=openpyxl.load_workbook(outfile)
                    workbook.sheetnames
                    sheet =workbook[query_number]
                    sheet.merge_cells(start_row=1, start_column=1, end_row=6, end_column=6)
                    sheet.merge_cells(start_row=1, start_column=1, end_row=6, end_column=6)
                    sheet.cell(row=1, column=1).value=str(query_comment)
                    workbook.save(outfile)
            except Exception as e:
                errFile = open("logs/env/error.log","w") # logs folder should be present in the current folder
                errFile.write(str(e))
                errFile.close()
                print("Could Not Run Query") 
                print(str(e))
                raise TypeError("Again !?!") 
                print(traceback.format_exc())

def run_insert_query(infile,verbose=False):
    with open(infile, 'r') as f:
        queries = [q.strip() for q in re.split(r';[\s\r\n]+', f.read()) if q.strip()]
    for i,query in enumerate(queries):
        query_list=re.split(r'-->', str(query))
        query_number=query_list[0]
        query_string=query_list[1]       
#         print("query_number="+query_number)
        if verbose:
            print('Running query: {}/{}\n{}'.format(i+1, len(queries), q))
        try:
            print("trying to run query\n\n"+query_string+"\n\n")
            _, _ = exec_teradata_sql(query_string)
        except Exception as e:
            errFile = open("logs/env/error.log","w")# logs folder should be present in the current folder
            errFile.write(str(e))
            errFile.close()
            print("Could Not Run Query") 
            print(str(e))
            raise TypeError("Again !?!") 
            print(traceback.format_exc())
            sys.exit(0)


def read_comment_file(comment_file):
    comments_od = OrderedDict()
    with open('comments.txt', 'r') as myfile:
        for line in myfile:
            print(line)
            a=re.split(r'=', str(line))
            comments_od[str(a[0])] = str(a[1]);
    return comments_od
 


def format_output(outfile):
# Delete default sheet
    columnList=['A','B','C','D','E','F','G','H','I','J']
    workbook=openpyxl.load_workbook(outfile)
    print(workbook.sheetnames)
    std=workbook['Sheet']
    print("deleting sheet"+str(std))
    workbook.remove(std)
    print(workbook.sheetnames)
# Set column width to 10 for all columns in all the sheets
    worksheet_names = workbook.sheetnames
    for sht in workbook.sheetnames:
        print(sht)
        sheet_index = worksheet_names.index(sht)
        workbook.active=sheet_index
        sheet=workbook.active
        for i in range(0,10):
            sheet.column_dimensions[columnList[i]].width=20
    try:
        workbook.save(outfile)
    except:
        print("Could Not write to workbook")
        errFile = open("logs/env/error.log","w")# logs folder should be present in the current folder
        errFile.write(str(e))
        errFile.close()


def replace_args(input_file,arg_file):
    print('\n\n\n\narg_file: '+arg_file)
    with open(input_file, 'r') as file:
        input_queries = file.read()
        print(input_queries)
        args_od = OrderedDict()
        with open(arg_file, 'r') as myfile:
            args_array = myfile.read().splitlines()
        print(args_array)
        final_queries = str(input_queries).format(*args_array)
    with open("Input_file_with_args.txt", "w") as text_file:
        text_file.write(final_queries)

def check_arg_file(argument_file):
    pst = pytz.timezone('America/Los_Angeles')
    if (argument_file == "NA") or (argument_file == ""):
        print("Print Current Date")
        curr_date=datetime.date(datetime.now(pst))
        with open("curr_date_run.txt", 'w') as myfile:
            myfile.write(str(curr_date))
        argument_file="curr_date_run.txt"
        return argument_file
    else:
        with open(argument_file, 'r') as myfile:
            args_array = myfile.readlines()
        if (len(args_array) == 0):
            print("Print Current Date")
            curr_date=datetime.date(datetime.now(pst))
            with open("curr_date_run.txt", 'w') as myfile:
                myfile.write(str(curr_date))
            argument_file="curr_date_run.txt"
            return argument_file
        else:
            return argument_file


# Main

print("Number of arguments: " +str(len(sys.argv))+ " arguments.")
print("Argument List:"+ str(sys.argv))
arguments=sys.argv

try:
    input_file=str(arguments[1])
    comment_file=str(arguments[2])
    arg_file=str(arguments[3])
    output_file=str(arguments[4])
    job_type=str(arguments[5])
except:
    print("Need total 5 arguments you gave "+str(len(sys.argv)-1)+". Please try again")
    errFile = open("logs/env/error.log","w")# logs folder should be present in the current folder
    errFile.write(str(e))
    errFile.close()
    sys.exit(0)
    

print("input_file: "+input_file)
print("comment_file: "+comment_file)
print("arg_file: "+arg_file)
print("output_file: "+output_file)
print("job_type: "+job_type)

arg_file=check_arg_file(arg_file)

input_file_with_args="input_file_with_args_"+job_type

if (job_type == "historical_load_daily") or (job_type == "historical_load_weekly") or (job_type == "historical_load_monthly"):
    with open(input_file, 'r') as file:
        input_queries = file.read()
    with open(arg_file, 'r') as myfile:
        args_array = myfile.read().splitlines()
    print("\n\nargs_array\n\n"+str(args_array[0])+"\n\n\n\n")
    start_date=datetime.datetime.strptime(args_array[0] , '%Y-%m-%d').date()
    start_date_1=start_date
    end_date=datetime.datetime.strptime(args_array[1] , '%Y-%m-%d').date()
    while(start_date_1 <= end_date):
        print("start_date_1 "+str(start_date_1))
        print("end_date "+str(end_date))
        final_queries = str(input_queries).format(str(start_date_1))
        print("final_queries "+str(final_queries))
        with open(input_file_with_args, "w") as text_file:
            text_file.write(final_queries)
        run_insert_query(input_file_with_args)
        if(job_type=="historical_load_daily"):
            start_date_1=(start_date_1 +datetime.timedelta(days=1))
        elif (job_type=="historical_load_weekly"):
            start_date_1=(start_date_1 +datetime.timedelta(weeks=1))
        elif (job_type=="historical_load_monthly"):
            start_date_1=(start_date_1 +relativedelta(months=+1))


elif(job_type=="insert"):
    replace_args(input_file,arg_file)
    run_insert_query('Input_file_with_args.txt')
    
elif(job_type=="select"):
    wb = openpyxl.Workbook()
    try:
        os.remove(output_file)
    except OSError:
        pass
    wb.save(output_file) 
    comments_od=read_comment_file(comment_file)
    replace_args(input_file,arg_file)
    run_query_file("Input_file_with_args.txt",output_file,comments_od)
    format_output(output_file)
else:
    print("Job Type can only be *hist, select or insert, you gave: "+job_type)

